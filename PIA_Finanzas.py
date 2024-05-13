#Repositorio: https://github.com/Angel-Diaz-H/PIA-Finanzas
#--------------------------------------MÓDULOS.
import os
import pandas as pd
import sqlite3
from sqlite3 import Error
import sys
from unidecode import unidecode
import re
from tabulate import tabulate
from colorama import Fore, Style
from datetime import datetime, timedelta
import datetime as dt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#--------------------------------------FUNCIONES DE IMPRESIÓN CON ESTILOS.
def aviso(mensaje, longitud):
    mensaje = mensaje.upper()
    print(f"\n{Style.BRIGHT}{guiones(longitud)} {mensaje} {guiones(longitud)}{Style.RESET_ALL}\n")

def avisoGreen(mensaje, longitud):
    mensaje = mensaje.upper()
    print(f"\n{Fore.GREEN}{Style.BRIGHT}{guiones(longitud)} {mensaje} {guiones(longitud)}{Style.RESET_ALL}\n")
    
def avisoRed(mensaje, longitud):
    mensaje = mensaje.upper()
    print(f"\n{Fore.RED}{Style.BRIGHT}{guiones(longitud)} {mensaje} {guiones(longitud)}{Style.RESET_ALL}\n")

def avisoBlue(mensaje, longitud):
    mensaje = mensaje.upper()
    print(f"\n{Fore.BLUE}{Style.BRIGHT}{guiones(longitud)} {mensaje} {guiones(longitud)}{Style.RESET_ALL}\n")


def printNegrita(mensaje):
    print(f"{Style.BRIGHT}{mensaje}{Style.RESET_ALL}")

def inputNegrita(mensaje):
    return input(f"{Style.BRIGHT}{mensaje}{Style.RESET_ALL}")


def printGreen(mensaje):
    print(f"{Fore.GREEN}{mensaje}{Style.RESET_ALL}")

def printRed(mensaje):
    print(f"{Fore.RED}{mensaje}{Style.RESET_ALL}")

def printBlue(mensaje):
    print(f"{Fore.BLUE}{mensaje}{Style.RESET_ALL}")

def printCyan(mensaje):
    print(f"{Fore.CYAN}{mensaje}{Style.RESET_ALL}")


def printGreenNegrita(mensaje):
    print(f"{Fore.GREEN}{Style.BRIGHT}{mensaje}{Style.RESET_ALL}")

def printRedNegrita(mensaje):
    print(f"{Fore.RED}{Style.BRIGHT}{mensaje}{Style.RESET_ALL}")

def printBlueNegrita(mensaje):
    print(f"{Fore.BLUE}{Style.BRIGHT}{mensaje}{Style.RESET_ALL}")

def printCyanNegrita(mensaje):
    print(f"{Fore.CYAN}{Style.BRIGHT}{mensaje}{Style.RESET_ALL}")


def inputGreen(mensaje):
    return input(f"{Fore.GREEN}{mensaje}{Style.RESET_ALL}")

def inputRed(mensaje):
    return input(f"{Fore.RED}{mensaje}{Style.RESET_ALL}")

def inputBlue(mensaje):
    return input(f"{Fore.BLUE}{mensaje}{Style.RESET_ALL}")

def inputCyan(mensaje):
    return input(f"{Fore.CYAN}{mensaje}{Style.RESET_ALL}")

def inputGreenNegrita(mensaje):
    return input(f"{Fore.GREEN}{Style.BRIGHT}{mensaje}{Style.RESET_ALL}")

def inputRedNegrita(mensaje):
    return input(f"{Fore.RED}{Style.BRIGHT}{mensaje}{Style.RESET_ALL}")

def inputBlueNegrita(mensaje):
    return input(f"{Fore.BLUE}{Style.BRIGHT}{mensaje}{Style.RESET_ALL}")

def inputCyanNegrita(mensaje):
    return input(f"{Fore.CYAN}{Style.BRIGHT}{mensaje}{Style.RESET_ALL}")
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#--------------------------------------FUNCIONES GENERALES.
def limpiar_consola():
    os.system('cls' if os.name == 'nt' else 'clear')

def guiones(longitud):
    return '-' * longitud

def respuestaSiYNo():
    while True:
        respuesta = darFormatoATexto(inputNegrita('\tRespuesta: '))
        if respuesta == 'SI':
            return True
        elif respuesta == 'NO':
            return False
        else:
            printRedNegrita('\n\tIngrese una respuesta válida (Sí/No).')

def darFormatoATexto(texto, sinEspacios = False):
    texto = unidecode(texto).strip().upper()
    if sinEspacios:
        texto = texto.replace(' ', '')
    else:
        texto = re.sub(r'\s+', ' ', texto)
    return texto

def validarSalir(texto):
    texto = darFormatoATexto(str(texto), True)
    if texto == '<REGRESAR>':
        return True
    return False

def indicarEnter():
    inputBlueNegrita("\n\nDe clic en Enter para continuar.")

def validarTextoValido(texto, permitirNumero = True):
    
    if not permitirNumero == True:
        if not texto.replace(' ', '').isalpha():
            printRedNegrita("Ingrese solo letras.\n")
            return True
    else:
        if not re.match("^[A-Za-z0-9]*$", texto.replace(' ', '')):
            printRedNegrita("Ingrese solo letras y números.\n")
            return True

    if len(texto) < 3 or len(texto) > 50:
        printRedNegrita("Ingrese un nombre válido.\n")
        return True
    return False

def mensajeInicialEnFuncionesEspecificas():
    printBlueNegrita("Para regresar al menú anterior ingrese:")
    printGreenNegrita("<Regresar>\n")

def solicitarEnteroOSalir(descripcion):
    while True:
        respuesta = inputNegrita(f"{descripcion}: ")

        if respuesta.isdigit():
            return int(respuesta), False
        elif validarSalir(respuesta):
            return True, True
        else:
            printRedNegrita("Ingrese un número válido.\n")

def solicitarEntero_Salir_Mostrar(descripcion):
    while True:
        respuesta = inputNegrita(f"{descripcion}: ")
        if respuesta.isdigit():
            return int(respuesta), False
        elif validarSalir(respuesta):
            return True, True
        elif darFormatoATexto(respuesta) == '<MOSTRAR>':
            return '<MOSTRAR>', False
        else:
            printRedNegrita("Ingrese un número válido.\n")

def solicitarRangoEnteroOSalir(opcionMin, opcionMax):
    while True:
        try:
            opcionValida = inputNegrita("Opción: ")
            opcionValida = int(opcionValida)
            if opcionValida >= opcionMin and opcionValida <= opcionMax:
                return opcionValida, False
            else:
                printRedNegrita(f"Ingrese un número válido ({opcionMin}-{opcionMax}).\n")
        except ValueError:
            if validarSalir(opcionValida):
                return True, True
            else:
                printRedNegrita(f"Ingrese un número válido ({opcionMin}-{opcionMax}).\n")
        except Error as e:
            inputRedNegrita(f'Se produjo el siguiente error: {e}')
        except Exception as e:
            inputRedNegrita(f'Se produjo el siguiente error: {sys.exc_info()[0]}')

def solicitarFechaOSalir(noMayorActual = False):
    while True:
        fecha = inputNegrita("Fecha: ").strip()
        try:
            fecha = dt.datetime.strptime(fecha, "%d/%m/%Y").date()
            if noMayorActual:
                if fecha > fechaActual():
                    printRedNegrita("Ingrese una fecha no mayor a hoy.\n")
                    continue
            return fecha, False
        except ValueError:
            if validarSalir(fecha):
                return None, True
            else:
                printRedNegrita("Ingrese una fecha válida.\n")   

def fechaActual():
    fecha_actual = dt.date.today()
    return fecha_actual

def solicitarFlotanteOSalir(descripcion):
    while True:
        flotante = inputNegrita(f"{descripcion}: ").strip()
        if re.match(r'^\d+(\.\d{1,2})?$', flotante):
            flotante = float(flotante)
            return flotante, False
        elif validarSalir(flotante):
            input("entro en el 2")
            return None, True
        else:
            printRedNegrita("Ingrese un número válido.\n")
            continue

#--------------------------------------FUNCIONES AUXILIARES. 
def mostrarOpcionesDeMenu(listaActual):
    print('Ingrese el número de la opción que desee realizar:')
    print(tabulate(listaActual, headers = 'firstrow', tablefmt = 'pretty'))

def validarOpcionesNumericas(opcionMin, opcionMax):
    while True:
        try:
            opcionValida = int(inputCyan("Opción: "))
            if opcionValida >= opcionMin and opcionValida <= opcionMax:
                return opcionValida
            else:
                printRedNegrita(f"Ingrese un número válido ({opcionMin}-{opcionMax}).\n")
        except ValueError:
            printRedNegrita(f"Ingrese un número válido ({opcionMin}-{opcionMax}).\n")

def contarCantidadOpcionesDeMenu(listaActual):
    cantidad = len(listaActual) - 1
    return cantidad

def mostrarYValidarMenu(ubicacion, opcion, lista):
    mostrarTitulo(ubicacion, True)  
    mostrarOpcionesDeMenu(lista)
    opcion = validarOpcionesNumericas(1, contarCantidadOpcionesDeMenu(lista))
    limpiar_consola() 
    if not opcion == contarCantidadOpcionesDeMenu(lista):
            ubicacion.append(lista[opcion][1])
    return opcion, ubicacion

def mostrarTitulo(ubicacion, esMenu = False):
    printNegrita(f'Ubicación: {" / ".join(ubicacion)}')
    if esMenu == True:
        avisoBlue(f'MENÚ {ubicacion[-1]}', 20)
    else:
        avisoBlue(f'{ubicacion[-1]}', 20)
        return
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#--------------------------------------IMPORTAR EL EXCEL.
aviso("Aviso:", 0)
try:
    df = pd.read_excel('Antiguedad_de_saldos.xlsx', engine = 'openpyxl')
    printGreenNegrita("El archivo Excel se ha leído correctamente.\n")
except Error as e:
        printRedNegrita(f'Se produjo el siguiente error: {e}')
except Exception:
        printRedNegrita(f'Se produjo el siguiente error: {sys.exc_info()[0]}')

#--------------------------------------CREACIÓN DE TABLAS SQL (SOLO SI NO EXISTEN).
def creacion_tablas():
    try:
        with sqlite3.connect('CuentasPorCobrar.db') as conn:
            mi_cursor = conn.cursor()
            mi_cursor.execute("CREATE TABLE IF NOT EXISTS Clientes \
                            (CLAVE_CLIENTE INTEGER PRIMARY KEY NOT NULL,\
                            NOMBRECLIENTE TEXT NOT NULL, \
                            ESTADOCLIENTE INTEGER NOT NULL);")
            mi_cursor.execute("CREATE TABLE IF NOT EXISTS CuentasPorCobrar \
                            (CLAVE_GUIA INTEGER PRIMARY KEY NOT NULL, \
                            DIASCARTERA INTEGER NOT NULL,\
                            FECHA timestamp NOT NULL, \
                            TOTAL INTEGER NOT NULL, \
                            CLAVE_CLIENTE INTEGER NOT NULL,\
                            FOREIGN KEY (CLAVE_CLIENTE) REFERENCES CLIENTES(CLAVE_CLIENTE));")
            printGreenNegrita("Las tablas SQL se han cargado correctamente\n")
    except Error as e:
        printRedNegrita(f'Se produjo el siguiente error: {e}')
    except Exception:
        printRedNegrita(f'Se produjo el siguiente error: {sys.exc_info()[0]}')
    finally:
        conn.close()

#--------------------------------------INSERTAR REGISTROS EXCEL A LA TABLA CLIENTES.
def insertar_clientes(df):
    try:
        with sqlite3.connect('CuentasPorCobrar.db') as conn:
            mi_cursor = conn.cursor()
            for i in df['Nombre'].unique():
                mi_cursor.execute("INSERT INTO Clientes (NOMBRECLIENTE, ESTADOCLIENTE) \
                                  SELECT ?, ? WHERE NOT EXISTS(SELECT 1 \
                                  FROM Clientes WHERE NOMBRECLIENTE = ?)", (i, 1, i))
            printGreenNegrita("Los datos se han insertado correctamente en la tabla Clientes.\n")
    except Error as e:
        printRedNegrita(f'Se produjo el siguiente error: {e}')
    except Exception:
        printRedNegrita(f'Se produjo el siguiente error: {sys.exc_info()[0]}')
    finally:
        conn.close()

#--------------------------------------INSERTAR REGISTROS EXCEL A LA TABLA CUENTASPORCOBRAR.
def insertar_cuentasporcobrar(df):
    try:
        with sqlite3.connect('CuentasPorCobrar.db') as conn:
            mi_cursor = conn.cursor()
            for i, row in df.iterrows():
                fecha = row['Fecha'].strftime('%Y-%m-%d %H:%M:%S')  # Convertir la fecha a una cadena de texto
                mi_cursor.execute("INSERT INTO CuentasPorCobrar (CLAVE_GUIA, DIASCARTERA, FECHA, TOTAL, CLAVE_CLIENTE) \
                                  SELECT ?, ?, ?, ?, ? WHERE NOT EXISTS(SELECT 1 FROM CuentasPorCobrar WHERE CLAVE_GUIA = ?) AND \
                                  EXISTS(SELECT 1 FROM Clientes WHERE CLAVE_CLIENTE = ?)", (row['Guía'], row['Días '], fecha, row['Total'], row['Cliente'], row['Guía'], row['Cliente']))
            printGreenNegrita("Los datos se han insertado correctamente en la tabla CuentasPorCobrar.\n")
    except Error as e:
        printRedNegrita(f'Se produjo el siguiente error: {e}')
    except Exception:
        printRedNegrita(f'Se produjo el siguiente error: {sys.exc_info()[0]}')
    finally:
        conn.close()

creacion_tablas()
insertar_clientes(df)
insertar_cuentasporcobrar(df)
indicarEnter()
limpiar_consola()
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#--------------------------------------LISTAS DE MENÚS.
lmenu_principal = [('Opción', 'Descripción'),
              (1, 'Clientes'),
              (2, 'Cuentas por cobrar'),
              (3, 'Salir')]
lmenu_clientes = [('Opción', 'Descripción'),
              (1, 'Registrar clientes'),
              (2, 'Editar nombre de clientes'),
              (3, 'Suspender clientes'),
              (4, 'Recuperar clientes'),
              (5, 'Mostrar clientes'),
              (6, 'Volver al menú principal')]
lmenu_cuentaPorCobrar = [('Opción', 'Descripción'),
              (1, 'Registrar cuentras por cobrar'),
              (2, 'Análisis de cuentas por cobrar'),
              (3, 'Volver al menú principal')]
lmenu_clientes_orden = [('Opción', 'Orden'),
              (1, 'Por clave'),
              (2, 'Por nombre'),
              (3, 'Por estado')]
lmenu_cuentaPorCobrar_diasDeCartera = [('Opción', 'Días de cartera', ''),
              (1, '15 días', 15),
              (2, '25 días', 25),
              (3, '30 días', 30),
              (4, '45 días', 45),
              (5, '60 días', 60)]
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#--------------------------------------1. MENÚ PRINCIPAL.
def menuPrincipal():
    opcion = 0

    while True:    
        ubicacion = ['Menú principal']
        if opcion == 0:
            avisoBlue('CUENTAS POR COBRAR', 20)
            mostrarOpcionesDeMenu(lmenu_principal)

            opcion = validarOpcionesNumericas(1, contarCantidadOpcionesDeMenu(lmenu_principal))

        if not opcion == contarCantidadOpcionesDeMenu(lmenu_principal):
            ubicacion.append(lmenu_principal[opcion][1])
            limpiar_consola()

        if opcion == 1:
            menuClientes(ubicacion)
        elif opcion == 2:
            menuCuentasPorCobrar(ubicacion)
        else:
            printCyanNegrita('\n¿Está seguro que desea salir? (Sí/No)')
            if respuestaSiYNo():
                avisoGreen("Archivo cerrado correctamente.", 25)
                avisoGreen("Gracias por usar nuestro sistema, hasta la próxima.", 15)
                break
            else:
                limpiar_consola()
                opcion = 0
        opcion = 0

#--------------------------------------1.1. MENÚ CLIENTES.
def menuClientes(ubicacion):
    ubicacionOriginal = ubicacion.copy()
    opcion = 0

    while True:    
        ubicacion = ubicacionOriginal.copy()
        if opcion == 0:
            opcion, ubicacion = mostrarYValidarMenu(ubicacion, opcion, lmenu_clientes)

        if opcion == 1:
            mostrarTitulo(ubicacion)
            registrarClientes()
        elif opcion == 2:
            mostrarTitulo(ubicacion)
            editarNombreClientes()
        elif opcion == 3:
            mostrarTitulo(ubicacion)
            suspenderCliente()
        elif opcion == 4:
            mostrarTitulo(ubicacion)
            recuperarClientes()
        elif opcion == 5:
            mostrarTitulo(ubicacion)
            mostrarClientes()
        else:
            break

        opcion = 0
        limpiar_consola()

#--------------------------------------1.2. MENÚ CUENTAS POR COBRAR.
def menuCuentasPorCobrar(ubicacion):
    ubicacionOriginal = ubicacion.copy()
    opcion = 0

    while True:    
        ubicacion = ubicacionOriginal.copy()
        if opcion == 0:
            opcion, ubicacion = mostrarYValidarMenu(ubicacion, opcion, lmenu_cuentaPorCobrar)

        if opcion == 1:
            mostrarTitulo(ubicacion)
            registrarCuentasPorCobrar()
        elif opcion == 2:
            mostrarTitulo(ubicacion)
            analisisDeCuentasPorCobrar()
        else:
            break

        opcion = 0
        limpiar_consola()
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#--------------------------------------1.1.1. OPCIÓN REGISTRAR CLIENTES.
def registrarClientes():
    mensajeInicialEnFuncionesEspecificas()
    print('Ingrese el nombre del cliente (empresa).')
    while True:
        nombre = darFormatoATexto(inputNegrita('Nombre: '))
        
        if validarSalir(nombre): break
        if validarTextoValido(nombre): continue

        try:
            with sqlite3.connect('CuentasPorCobrar.db') as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute("SELECT 1 FROM Clientes WHERE NOMBRECLIENTE = ?", (nombre,))
                existencia = mi_cursor.fetchone()
                if existencia: 
                    printRedNegrita("La empresa ya se encuentra registrada.\n")
                else:
                    mi_cursor.execute("INSERT INTO Clientes(NOMBRECLIENTE, ESTADOCLIENTE) VALUES (?, ?)", (nombre, 1))
                    printCyanNegrita("\n¿Desea registrar el cliente? Confirme su respuesta (Sí/No).")
                    if respuestaSiYNo():
                        printGreenNegrita("\nRegistro de cliente exitosamente.")
                    else:
                        printBlueNegrita("\nEl cliente no se registró.")
                    indicarEnter()
                    break
        except Error as e:
            printRedNegrita(f'Se produjo el siguiente error: {e}')
        except Exception as e:
            printRedNegrita(f'Se produjo el siguiente error: {sys.exc_info()[0]}')
        finally:
            conn.close()

#--------------------------------------1.1.2. OPCIÓN EDITAR CLIENTES.
def editarNombreClientes():
    while True:
        try:
            with sqlite3.connect('CuentasPorCobrar.db') as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute("SELECT CLAVE_CLIENTE, NOMBRECLIENTE, ESTADOCLIENTE FROM Clientes")
                existencia = mi_cursor.fetchall()
                if existencia:
                    mensajeInicialEnFuncionesEspecificas()
                    existenciaImpresa = [(clave, nombre, "Activo" if estado else "Inactivo") for clave, nombre, estado in existencia]
                    printNegrita("Lista de clientes:")
                    print(tabulate(existenciaImpresa, headers = ["Clave del cliente", "Nombre", "Estado "], tablefmt = 'pretty'))
                    print("\nIngrese la clave del cliente a editar.")

                    while True:
                        respuesta = solicitarEnteroOSalir("Clave")
                        if respuesta[1]: break
                        busqueda = any(clave[0] == respuesta[0] for clave in existencia)
                        if busqueda:
                            print("\nIngrese el nuevo nombre de la empresa.")
                            nuevoNombre = darFormatoATexto(inputNegrita('Nuevo nombre: '))
                            if validarSalir(nuevoNombre): break
                            if validarTextoValido(nuevoNombre): continue

                            printCyanNegrita("\n¿Está seguro de actualizar el nombre? (Sí/No).")
                            if respuestaSiYNo():
                                mi_cursor.execute("UPDATE Clientes SET NOMBRECLIENTE = ? WHERE CLAVE_CLIENTE = ?", (nuevoNombre, respuesta[0]))
                                printGreenNegrita("\nActualización del nombre realizado exitosamente.")
                            else:
                                printBlueNegrita("El nombre del cliente no se actualizó.")
                            indicarEnter()
                            break
                        else:
                            printRedNegrita("Ingrese una clave válida o <regresar> para volver al menú anterior.\n")
                else:
                    printBlueNegrita('\nActualmente no se cuenta con clientes registrados.')
                    indicarEnter()
                    break
                break
        except Error as e:
            inputRedNegrita(f'Se produjo el siguiente error: {e}')
        except Exception as e:
            inputRedNegrita(f'Se produjo el siguiente error: {sys.exc_info()[0]}')
        finally:
            conn.close()

#--------------------------------------1.1.3. OPCIÓN SUSPENDER CLIENTES.
def suspenderCliente():
    while True:
        try:
            with sqlite3.connect('CuentasPorCobrar.db') as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute("SELECT CLAVE_CLIENTE, NOMBRECLIENTE FROM Clientes WHERE ESTADOCLIENTE = 1")
                existencia = mi_cursor.fetchall()
                if existencia:
                    mensajeInicialEnFuncionesEspecificas()
                    printNegrita("Lista de clientes activos:")
                    print(tabulate(existencia, headers = ["Clave del cliente", "Nombre"], tablefmt = 'pretty'))
                    print("\nIngrese la clave del cliente a suspender.")

                    while True:
                        respuesta = solicitarEnteroOSalir("Clave")
                        if respuesta[1]: break
                        busqueda = any(clave[0] == respuesta[0] for clave in existencia)
                        if busqueda:
                            printCyanNegrita("\n¿Desea suspender el siguiente cliente? (Sí/No)")
                            mi_cursor.execute("SELECT CLAVE_CLIENTE, NOMBRECLIENTE FROM Clientes WHERE CLAVE_CLIENTE = ?", (respuesta[0],))
                            clienteEncontrado = mi_cursor.fetchone()
                            print(tabulate([list(clienteEncontrado)], headers = ["Clave del cliente", "Nombre"], tablefmt = 'pretty'))

                            if respuestaSiYNo():
                                mi_cursor.execute("UPDATE Clientes SET ESTADOCLIENTE = 0 WHERE CLAVE_CLIENTE = ?", (clienteEncontrado[0],))
                                printGreenNegrita("\nCliente suspendido con éxito.")
                            else:
                                printBlueNegrita("\nEl cliente no se suspendió.")
                            indicarEnter()
                            break
                        else:
                            printRedNegrita("Ingrese una clave válida o <regresar> para volver al menú anterior.\n")
                else:
                    printBlueNegrita('\nActualmente no se cuenta con clientes activos.')
                    indicarEnter()
                    break
                break
        except Error as e:
            inputRedNegrita(f'Se produjo el siguiente error: {e}')
        except Exception as e:
            inputRedNegrita(f'Se produjo el siguiente error: {sys.exc_info()[0]}')
        finally:
            conn.close()

#--------------------------------------1.1.4. OPCIÓN RECUPERAR CLIENTES.
def recuperarClientes():
    while True:
        try:
            with sqlite3.connect('CuentasPorCobrar.db') as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute("SELECT CLAVE_CLIENTE, NOMBRECLIENTE FROM Clientes WHERE ESTADOCLIENTE = 0")
                existencia = mi_cursor.fetchall()
                if existencia:
                    mensajeInicialEnFuncionesEspecificas()
                    printNegrita("Lista de clientes suspendidos:")
                    print(tabulate(existencia, headers = ["Clave del cliente", "Nombre"], tablefmt = 'pretty'))
                    print("\nIngrese la clave del cliente a recuperar.")

                    while True:
                        respuesta = solicitarEnteroOSalir("Clave")
                        if respuesta[1]: break
                        busqueda = any(clave[0] == respuesta[0] for clave in existencia)
                        if busqueda:
                            printCyanNegrita("\n¿Desea recuperar el siguiente cliente? (Sí/No)")
                            mi_cursor.execute("SELECT CLAVE_CLIENTE, NOMBRECLIENTE FROM Clientes WHERE CLAVE_CLIENTE = ?", (respuesta[0],))
                            clienteEncontrado = mi_cursor.fetchone()
                            print(tabulate([list(clienteEncontrado)], headers = ["Clave del cliente", "Nombre"], tablefmt = 'pretty'))

                            if respuestaSiYNo():
                                mi_cursor.execute("UPDATE Clientes SET ESTADOCLIENTE = 1 WHERE CLAVE_CLIENTE = ?", (clienteEncontrado[0],))
                                printGreenNegrita("\nCliente recuperado con éxito.")
                            else:
                                printBlueNegrita("\nEl cliente no se recuperó.")
                            indicarEnter()
                            break
                        else:
                            printRedNegrita("Ingrese una clave válida o <regresar> para volver al menú anterior.\n")
                else:
                    printBlueNegrita('\nActualmente no se cuenta con clientes suspendidos.')
                    indicarEnter()
                    break
                break
        except Error as e:
            inputRedNegrita(f'Se produjo el siguiente error: {e}')
        except Exception as e:
            inputRedNegrita(f'Se produjo el siguiente error: {sys.exc_info()[0]}')
        finally:
            conn.close()

#--------------------------------------1.1.5. OPCIÓN MOSTRAR CLIENTES.
def mostrarClientes():
    while True:
        try:
            with sqlite3.connect('CuentasPorCobrar.db') as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute("SELECT CLAVE_CLIENTE, NOMBRECLIENTE, ESTADOCLIENTE FROM Clientes")
                existencia = mi_cursor.fetchall()
                if existencia:
                    mensajeInicialEnFuncionesEspecificas()
                    existenciaImpresa = [(clave, nombre, "Activo" if estado else "Inactivo") for clave, nombre, estado in existencia]
                    mostrarOpcionesDeMenu(lmenu_clientes_orden)
                    respuesta = solicitarRangoEnteroOSalir(1, 3)
                    
                    if respuesta[1]: break
                    printNegrita("\nLista de clientes:")
                    if respuesta[0] == 1:
                        existenciaImpresaOrdenada = sorted(existenciaImpresa, key = lambda x: x[0])
                        print(tabulate(existenciaImpresaOrdenada, headers = ['Clave', 'Nombre', 'Estado'], tablefmt = 'pretty'))
                    elif respuesta[0] == 2:
                        existenciaImpresaOrdenada = sorted(existenciaImpresa, key = lambda x: x[1])
                        print(tabulate(existenciaImpresaOrdenada, headers = ['Clave', 'Nombre', 'Estado'], tablefmt = 'pretty'))
                    else:
                        existenciaImpresaOrdenada = sorted(existenciaImpresa, key = lambda x: x[2])
                        print(tabulate(existenciaImpresaOrdenada, headers = ['Clave', 'Nombre', 'Estado'], tablefmt = 'pretty'))
                    
                    printCyanNegrita("\n¿Desea exportar la información a Excel? (Sí/No)")
                    if respuestaSiYNo():
                        fechaHora = datetime.now().strftime("%d-%m-%Y_%H%M%S")
                        nombre_archivo = f"ListadoClientes_{fechaHora}.xlsx"
                        df = pd.DataFrame(existenciaImpresaOrdenada, columns = ["Clave", "Nombre", "Estado cliente"])
                        df.to_excel(nombre_archivo, index = False)
                        printGreenNegrita("\nInformación exportada exitosamente a Excel.")
                        printBlueNegrita(f"Nombre del archivo:")
                        printNegrita(f"{nombre_archivo}")
                        indicarEnter()
                else:
                    printBlueNegrita('\nActualmente no se cuenta con clientes registrados.')
                    indicarEnter()
                break
        except Error as e:
            inputRedNegrita(f'Se produjo el siguiente error: {e}')
        except Exception as e:
            inputRedNegrita(f'Se produjo el siguiente error: {sys.exc_info()[0]}')
        finally:
            conn.close()
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#
#--------------------------------------1.2.1. OPCIÓN REGISTRAR CUENTAS POR COBRAR.
def registrarCuentasPorCobrar():
    while True:
        try:
            with sqlite3.connect('CuentasPorCobrar.db',
                             detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                mi_cursor = conn.cursor()
                mi_cursor.execute("SELECT CLAVE_CLIENTE, NOMBRECLIENTE FROM Clientes WHERE ESTADOCLIENTE = 1")
                existencia = mi_cursor.fetchall()
                
                if not existencia:
                    printBlueNegrita('\nActualmente no se cuenta con clientes registrados o activos.')
                    indicarEnter()
                    break

                mensajeInicialEnFuncionesEspecificas()
                printBlueNegrita("Para ver los clientes disponibles ingrese:")
                printGreenNegrita("<Mostrar>")
                
                #Solicitar clave del cliente.
                continuarPrograma = None
                print("\nIngrese la clave del cliente a registrar cuenta por cobrar.")
                while True:
                    respuesta = solicitarEntero_Salir_Mostrar("Clave")
                    if respuesta[1]: break
                    if respuesta[0] == '<MOSTRAR>': 
                        printNegrita("\nLista de clientes:")
                        print(tabulate(existencia, headers = ["Clave del cliente", "Nombre"], tablefmt = 'pretty'))
                        continue     
                    confirmarExistencia = any(clave[0] == respuesta[0] for clave in existencia)
                    if not confirmarExistencia:
                        printRedNegrita("Ingrese una clave válida o <regresar> para volver al menú anterior.\n") 
                        continue
                    for tupla in existencia:
                        if tupla[0] == respuesta[0]: claveYNombreCliente = tupla
                    continuarPrograma = True
                    break
                
                if not continuarPrograma: break
                continuarPrograma = None
                print('')

                #Solicitar días de cartera
                mostrarOpcionesDeMenu(lmenu_cuentaPorCobrar_diasDeCartera)
                respuesta = solicitarRangoEnteroOSalir(1, 5)
                if respuesta[1]: break
                for diasCartera in lmenu_cuentaPorCobrar_diasDeCartera:
                    if diasCartera[0] == respuesta[0]:
                        diaCartera = diasCartera[2] 
                        break
                
                #Solicitar fecha
                print("\nIngrese la fecha de la cuenta por cobrar (dd/mm/aaaa).")
                fecha = solicitarFechaOSalir(True)
                if fecha[1]: break
                fechaImpresa = fecha[0].strftime("%d-%m-%Y")

                #Solicitar total
                print("\nIngrese el total de la cuenta por cobrar (Máximo dos decimales).")
                total = solicitarFlotanteOSalir('Total')
                if total[1]: break

                tuplaImpresaCuentasPorCobrar = (claveYNombreCliente[0], claveYNombreCliente[1], diaCartera, fechaImpresa, total[0])
                tuplaCuentasPorCobrar = (claveYNombreCliente[0], diaCartera, fecha[0], total[0])
                printCyanNegrita("\n¿Desea registrar la cuenta por cobrar? Confirme su respuesta (Sí/No).")
                print(tabulate([list(tuplaImpresaCuentasPorCobrar)], headers = ['Clave del cliente', 'Nombre del cliente', 'Días de cartera', 'Fecha', 'Total'], tablefmt = 'pretty'))
                
                if respuestaSiYNo():
                    mi_cursor.execute('INSERT INTO CuentasPorCobrar(CLAVE_CLIENTE, DIASCARTERA, FECHA, TOTAL) VALUES \
                                      (?, ?, ?, ?)', tuplaCuentasPorCobrar)
                    printGreenNegrita("\nRegistro de cuenta por cobrar existoso")
                else:
                    printBlueNegrita("\nLa cuenta por cobrar no se registró.")
                indicarEnter()
                break
        except Error as e:
            inputRedNegrita(f'Se produjo el siguiente error: {e}')
        except Exception as e:
            inputRedNegrita(f'Se produjo el siguiente error: {sys.exc_info()[0]}')
        finally:
            conn.close()

#--------------------------------------1.2.6. OPCIÓN ANÁLISIS DE CUENTAS POR COBRAR.
def analisisDeCuentasPorCobrar():
    while True:
        try:
            with sqlite3.connect('CuentasPorCobrar.db',
                             detect_types = sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conn:
                mi_cursor = conn.cursor()

                #Validar existencia de cuentas por cobrar registradas.
                mi_cursor.execute("SELECT cpc.CLAVE_GUIA, cpc.DIASCARTERA, strftime('%d/%m/%Y', cpc.FECHA) AS 'FECHA', cpc.TOTAL, cpc.CLAVE_CLIENTE, c.NOMBRECLIENTE \
                                  FROM CuentasPorCobrar AS cpc JOIN Clientes AS c \
                                  ON cpc.CLAVE_CLIENTE = c.CLAVE_CLIENTE")
                existencia = mi_cursor.fetchall()
                if not existencia:
                    printBlueNegrita('\nActualmente no se cuenta con cuentas por cobrar registradas o activas.')
                    indicarEnter()
                    break

                mensajeInicialEnFuncionesEspecificas()

                #######################################################################33#ANTIGUEDAD DE SALDOS.
                #Solicitar fecha de antigüedad.
                print("Ingrese la fecha de antigüedad de saldos (dd/mm/aaaa).")
                fechaAntiguedad = solicitarFechaOSalir(True)
                if fechaAntiguedad[1]: break
                fechaAntiguedad = fechaAntiguedad[0]

                #Tupla existencia a lista.
                lExistencias = [list(tupla) for tupla in existencia]
                listaConFechaVencimiento = [lista + [(datetime.strptime(lista[2], '%d/%m/%Y') + timedelta(days=lista[1])).strftime('%d/%m/%Y')] for lista in lExistencias]
                listaConDiasVencidos = [lista + [max((fechaAntiguedad - datetime.strptime(lista[6], '%d/%m/%Y').date()).days, 0)] for lista in listaConFechaVencimiento]

                for lista in listaConDiasVencidos:
                    dias_vencidos = lista[7]
                    if dias_vencidos == 0:
                        lista.extend([lista[3], '', '', '', '', ''])
                    elif 0 < dias_vencidos <= 29:
                        lista.extend(['', lista[3], '', '', '', ''])
                    elif 30 <= dias_vencidos <= 59:
                        lista.extend(['', '', lista[3], '', '', ''])
                    elif 60 <= dias_vencidos <= 89:
                        lista.extend(['', '', '', lista[3], '', ''])
                    elif 90 <= dias_vencidos <= 179:
                        lista.extend(['', '', '', '', lista[3], ''])
                    elif dias_vencidos >= 180:
                        lista.extend(['', '', '', '', '', lista[3]])

                listaReordenada = [[lista[4], lista[5]] + lista[:4] + lista[6:] for lista in listaConDiasVencidos]

                #Sumatoria y porcentaje.
                suma_total = round(sum(lista[5] for lista in listaReordenada), 2)
                suma_al_corriente = round(sum(float(item[8]) if isinstance(item[8], (int, float)) else 0 for item in listaReordenada), 2)
                suma_menor_30 = round(sum(float(item[9]) if isinstance(item[9], (int, float)) else 0 for item in listaReordenada), 2)
                suma_menor_60 = round(sum(float(item[10]) if isinstance(item[10], (int, float)) else 0 for item in listaReordenada), 2)
                suma_menor_90 = round(sum(float(item[11]) if isinstance(item[11], (int, float)) else 0 for item in listaReordenada), 2)
                suma_menor_180 = round(sum(float(item[12]) if isinstance(item[12], (int, float)) else 0 for item in listaReordenada), 2)
                suma_mayor_180 = round(sum(float(item[13]) if isinstance(item[13], (int, float)) else 0 for item in listaReordenada), 2)
                sumatoria = ['', '', '', '', '', suma_total, '', '', suma_al_corriente, suma_menor_30, suma_menor_60, suma_menor_90, suma_menor_180, suma_mayor_180]

                pSumaAlCorriente = round(((suma_al_corriente / suma_total) * 100), 2)
                pSumaMenor30 = round(((suma_menor_30 / suma_total) * 100), 2)
                pSumaMenor60 = round(((suma_menor_60 / suma_total) * 100), 2)
                pSumaMenor90 = round(((suma_menor_90 / suma_total) * 100), 2)
                pSumaMenor180 = round(((suma_menor_180 / suma_total) * 100), 2)
                pSumaMayor180 = round(((suma_mayor_180 / suma_total) * 100), 2)
                pSumatoria = ['', '', '', '', '', '', '', '', pSumaAlCorriente, pSumaMenor30, pSumaMenor60, pSumaMenor90, pSumaMenor180, pSumaMayor180]
                
                listaReordenada.append(sumatoria)
                listaReordenada.append(pSumatoria)

                #Visualizar parte uno.
                printGreenNegrita("\nAnálisis de antigüedad de saldos realizado con éxito.")
                printCyanNegrita("¿Desea visualizarlo? (Sí/No)")
                if respuestaSiYNo():
                    printBlueNegrita("\nAntiguedad de saldos:")
                    print(tabulate(listaReordenada, headers = ['Clave cliente', 'Nombre cliente', 'Guía', 'Días Cartera', 'Fecha', 'Total', 'Fecha de vencimiento', 'Días vencidos', 'Al corriente', '< a 30', '< a 60', '< a 90', '< a 180', '> a 180'], tablefmt = 'pretty'))


                ###########################################################################REPORTE CORPORATIVO.
                #Crear el DataFrame
                df = pd.DataFrame(listaReordenada, columns = ['Clave cliente', 'Nombre cliente', 'Guía', 'Días Cartera', 'Fecha', 'Total', 'Fecha de vencimiento', 'Días vencidos', 'Al corriente', '< a 30', '< a 60', '< a 90', '< a 180', '> a 180'])
                df = df[['Nombre cliente', 'Al corriente', '< a 30', '< a 60', '< a 90', '< a 180', '> a 180']]

                #Eliminar filas donde 'Nombre cliente' es un espacio en blanco
                df = df[df['Nombre cliente'].str.strip() != '']

                for columna in ['Al corriente', '< a 30', '< a 60', '< a 90', '< a 180', '> a 180']:
                    #Verificar si los valores son numéricos
                    df[columna] = pd.to_numeric(df[columna], errors='coerce')

                #Agrupar y sumar
                df_agrupado = df.groupby('Nombre cliente').sum()

                #Calcular totales
                totales = df_agrupado.sum().round(2)
                df_totales = pd.DataFrame(totales).T
                df_totales.index = ['Total']
                df_final = pd.concat([df_agrupado, df_totales])

                # Calcular la suma de todos los totales
                suma_totales = df_final.loc['Total'].sum()

                # Calcular el porcentaje de cada total
                porcentajes = (df_final.loc['Total'] / suma_totales * 100).round(2)

                # Crear un nuevo DataFrame con los porcentajes
                df_porcentajes = pd.DataFrame(porcentajes).T
                df_porcentajes.index = ['Porcentaje']

                # Concatenar el DataFrame original con el DataFrame de porcentajes
                df_final_con_porcentajes = pd.concat([df_final, df_porcentajes])

                # Crear una copia del DataFrame para evitar modificar el original
                df_copia = df_final_con_porcentajes.copy()

                # Clasificar a los clientes
                df_copia.loc[df_copia['> a 180'] > 0, 'Categoría'] = 'Incobrable'
                df_copia.loc[(df_copia['< a 180'] > 0) & (df_copia['Categoría'].isna()), 'Categoría'] = 'Malo'
                df_copia.loc[(df_copia['< a 90'] > 0) & (df_copia['Categoría'].isna()), 'Categoría'] = 'Malo'
                df_copia.loc[(df_copia['< a 60'] > 0) & (df_copia['Categoría'].isna()), 'Categoría'] = 'Regular'
                df_copia.loc[(df_copia['< a 30'] > 0) & (df_copia['Categoría'].isna()), 'Categoría'] = 'Bueno'
                df_copia.loc[df_copia['Categoría'].isna(), 'Categoría'] = 'Bueno'

                # Eliminar la columna 'Categoría' para 'Total' y 'Porcentaje'
                df_copia.loc['Total', 'Categoría'] = ''
                df_copia.loc['Porcentaje', 'Categoría'] = ''

                # Convertir el DataFrame a lista incluyendo el índice
                lista_final_con_porcentajes_y_categoria = df_copia.reset_index().values.tolist()

                ###########################################################################INGLÉS
                # Crear un diccionario para mapear los encabezados al inglés
                traduccion = {
                    'Nombre cliente': 'Customer Name',
                    'Al corriente': 'Current',
                    '< a 30': '< 30',
                    '< a 60': '< 60',
                    '< a 90': '< 90',
                    '< a 180': '< 180',
                    '> a 180': '> 180',
                    'Categoría': 'Category'
                }

                # Crear un diccionario para mapear los valores de 'Categoría' al inglés
                traduccion_categoria = {
                    'Bueno': 'Good',
                    'Regular': 'Regular',
                    'Malo': 'Bad',
                    'Incobrable': 'Uncollectible',
                    '': ''  # Para manejar los valores vacíos
                }

                # Crear un diccionario para mapear 'Total' y 'Porcentaje' al inglés
                traduccion_indice = {
                    'Total': 'Total',
                    'Porcentaje': 'Percentage'
                }

                # Crear una copia en inglés
                df_copia_ingles = df_copia.rename(columns=traduccion)

                # Traducir los valores de 'Category'
                df_copia_ingles['Category'] = df_copia_ingles['Category'].map(traduccion_categoria)

                # Traducir 'Total' y 'Porcentaje' en el índice
                df_copia_ingles = df_copia_ingles.rename(index=traduccion_indice)

                # Convertir el DataFrame a lista incluyendo el índice
                lista_final_con_porcentajes_y_categoria_ingles = df_copia_ingles.reset_index().values.tolist()

                ###########################################################################IMPRESIÓN
                printGreenNegrita("\nAnálisis de reporte corporativo realizado con éxito.")
                printCyanNegrita("¿Desea visualizarlo? (Sí/No)")
                if respuestaSiYNo():
                    printBlueNegrita("\nReporte corporativo:")
                    print(tabulate(lista_final_con_porcentajes_y_categoria, headers = ['Nombre cliente', 'Al corriente', '< a 30', '< a 60', '< a 90', '< a 180', '> a 180', 'Categoría'], tablefmt = 'pretty'))
                    print("")

                    # Cuentas incobrables
                    cuentas_incobrables_monto = df_copia_ingles.loc['Total', '< 90'] + df_copia_ingles.loc['Total', '< 180'] + df_copia_ingles.loc['Total', '> 180']
                    total_sin_categoria = df_copia_ingles.drop(columns='Category').loc['Total'].sum()
                    cuentas_incobrables_porcentaje = (cuentas_incobrables_monto / total_sin_categoria * 100).round(2)
                    printBlueNegrita("Estimación de cuentas incobrables:")
                    print(f"${cuentas_incobrables_monto} ({cuentas_incobrables_porcentaje}%).")

                    # Mejor cliente
                    clientes_buenos = df_copia_ingles[df_copia_ingles['Category'] == 'Good']
                    clientes_buenos_sin_categoria = clientes_buenos.drop(columns='Category')
                    mejor_cliente = clientes_buenos_sin_categoria.sum(axis=1).idxmin()
                    printBlueNegrita("\nEl mejor cliente:")
                    print(f'{mejor_cliente}.')

                    # Clientes que dejaron de comprar
                    clientes_dejaron_comprar = df_copia_ingles[(df_copia_ingles['> 180'] > 0) & (df_copia_ingles.index != 'Total') & (df_copia_ingles.index != 'Percentage')]
                    nombres_clientes_dejaron_comprar = clientes_dejaron_comprar.index.tolist()
                    printBlueNegrita("\nClientes que dejaron de comprar:")
                    nombres = [[nombre] for nombre in nombres_clientes_dejaron_comprar]
                    print(tabulate(nombres, headers=['Nombres'], tablefmt='pretty'))


                    printBlueNegrita("\n\n\nCorporate reporting:")
                    print(tabulate(lista_final_con_porcentajes_y_categoria_ingles, headers = ['Customer Name', 'Current', '< 30', '< 60', '< 90', '< 180', '> 180', 'Category'], tablefmt = 'pretty'))
                    printBlueNegrita("Allowance for doubtful accounts:")
                    print(f"${cuentas_incobrables_monto} ({cuentas_incobrables_porcentaje}%)")
                    printBlueNegrita("\nThe best customer:")
                    print(f'{mejor_cliente}.')
                    printBlueNegrita("\nCustomers who stopped buying:")
                    print(tabulate(nombres, headers=['Names'], tablefmt='pretty'))


                ###########################################################################EXCEL.
                #Exportar parte uno.
                printCyanNegrita("\n¿Desea exportarlo a Excel? (Sí/No)")
                if respuestaSiYNo():
                    df = pd.DataFrame(listaReordenada, columns = ['Clave cliente', 'Nombre cliente', 'Guía', 'Días Cartera', 'Fecha', 'Total', 'Fecha de vencimiento', 'Días vencidos', 'Al corriente', '< a 30', '< a 60', '< a 90', '< a 180', '> a 180'])
                    fechaHora = datetime.now().strftime("%d-%m-%Y_%H%M%S")
                    nombre_archivo = f"Análisis_cuentas_por_cobrar_{fechaHora}.xlsx"
                    df.to_excel(nombre_archivo, index=False)
                    libro = load_workbook(nombre_archivo)
                    hoja = libro.active
                    # Cambiar el nombre de la hoja
                    hoja.title = 'Antigüedad de saldos'

                    for i, columna in enumerate(hoja.columns, start=1):
                        max_length = 0
                        columna = [str(celda.value) for celda in columna]
                        for celda in columna:
                            try:
                                if len(celda) > max_length: max_length = len(celda)
                            except:
                                pass
                        ajuste_ancho = (max_length + 2)
                        hoja.column_dimensions[get_column_letter(i)].width = ajuste_ancho
                    tab_range = "A1:N" + str(hoja.max_row)
                    tabla = Table(displayName="Tabla", ref=tab_range)
                    hoja.add_table(tabla) 
                    libro.save(nombre_archivo)

                    # Crear un nuevo escritor de Excel con pandas
                    with pd.ExcelWriter(nombre_archivo, engine='openpyxl', mode='a') as writer:
                        # Exportar el DataFrame en español a la segunda hoja
                        df_copia.to_excel(writer, sheet_name='Español')

                        # Exportar el DataFrame en inglés a la tercera hoja
                        df_copia_ingles.to_excel(writer, sheet_name='Inglés')
                    del writer

                    printGreenNegrita("\nInformación exportada exitosamente a Excel.")
                    printBlueNegrita(f"Nombre del archivo:")
                    printNegrita(f"{nombre_archivo}")   
                indicarEnter()
                break
        except Error as e:
            inputRedNegrita(f'Se produjo el siguiente error: {e}')
        except Exception as e:
            inputRedNegrita(f'Se produjo el siguiente error: {sys.exc_info()[0]}')
        finally:
            conn.close()
menuPrincipal()
